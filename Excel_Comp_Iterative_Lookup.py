#!/usr/bin/env python
# -*- coding: utf-8 -*-

## ADAPATED VERSION FOR ROW by ROW and then CELL by CELL Comparison ##

import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl import load_workbook
import csv
import re
from string import punctuation


ready = 'Helix_Case_PY.xlsx'

sheet_list_old = ['Helix_Case_DCW']
sheet_list_new = ['Helix_Case_AUDIT']

  
    
#function to rerun lookup less one attribute
def list_stripper(row, n):

    row_list_cat = []
    
    try:
        [row_list_cat.append([''.join([str(cell.internal_value).strip() for cell in row[0:n]]), cell.coordinate, n])]
        
    #handler for empty rows
    except TypeError, AttributeError:
        row = ['blank_cell']
        [row_list_cat.append(['blank_cell', cell.coordinate, n])]
    
    
    return row_list_cat
        
    
#function to write CSV file
def writer_csv(output_list, sheetname):
    
    #uses group name from URL to construct output file name
    file_out = "DCW_Compare_{dcw}.csv".format(dcw = sheetname.rsplit('.',2)[0])
    
    with open(file_out, 'w') as csvfile:
        col_labels = ['Compare_ID', 'Columns_Compared', 'Lookup_String', 'DCW_CellRef', 'Closest_Match_Audit']
        
        writer = csv.writer(csvfile, lineterminator='\n', delimiter=',', quotechar='"')
        newrow = col_labels
        writer.writerow(newrow)
        
        for i in output_list:
            
            newrow = i['compare_id'], i['columns_compared'], i['lookup_value_DCW'], i['cell_ref_dcw'], i['lookup_value_AUDIT']
            writer.writerow(newrow)      

    
#iterate through sheets and identify cells that do not match 
def sheet_checker(ready):    
    

    #load workbooks for DCW and Audit Report
    wb_all = openpyxl.load_workbook(ready, use_iterators=True, data_only=True)

 
    for i, j in map(None, sheet_list_new, sheet_list_old):

        ws_old = wb_all.get_sheet_by_name(j)
        ws_new = wb_all.get_sheet_by_name(i)
        
        row_old_list = []
        row_new_list = []
                
        compare_old_list = []
        compare_new_list = []
        
        output_list = []
        
        #build list from DCW
        for row_old in ws_old.iter_rows():
            col_count = len(row_old)
            
            if row_old is not None:
                
                n_old = len(row_old) + 1
            
                #create lists of 'cascading' concatenations
                for n in range(1, n_old):

                    compare_old = list_stripper(row_old, n)
                    compare_old_list.append(compare_old)
            
        #build list from AUDIT
        for row_new in ws_new.iter_rows():
            col_count = len(row_new)
            
            if row_new is not None:
                
                n_new = len(row_new) + 1
            
                #create lists of 'cascading' concatenations
                for n in range(1, n_new):

                    compare_new = list_stripper(row_new, n)
                    compare_new_list.append(compare_new)
            
          
        #create list for only concatenated strings from NEW
        list_lookup_new = []
        for p in compare_new_list:
            for q in p:
                list_lookup_new.append(q[0].strip())
        
        #loop through DCW and see if items cells within rows match entries in AUDIT
        x=1         #initialise counter
        #prochar = re.compile('[(=\-\+\:/&<>;|\'"\?%#$@\,\._)]')
        prochar = re.compile('.-')

        
        for e in compare_old_list:
            mismatch_dict = {}
            for f in e:
                if f[0].strip() not in list_lookup_new:
                    
                    #regex pattern to find closest match          
                    #text = f[0].strip()
                    #text = prochar.sub(' ', f[0]).strip()
                    text = f[0].replace('-', ' ').replace('.', ' ').strip()
                    #pattern = re.compile(r'[^-.]({text})*', re.IGNORECASE)
                    pattern = re.compile(text, re.IGNORECASE)

                    
                    #build dictionary of items
                    mismatch_dict = {
                        'compare_id' : str(x),
                        'columns_compared' : str(e[0][2]),
                        'lookup_value_DCW' : str(e[0][0].strip()),
                        #'lookup_value_AUDIT' : ', '.join(set(filter(None, [pattern.search(z.strip()).group() if pattern.search(z.strip()) is not None else "" for z in list_lookup_new]))),
                        #'lookup_value_AUDIT' : ', '.join(set(filter(None, [z if pattern.search(prochar.sub(' ', z).strip().strip()) is not None else "" for z in list_lookup_new]))),
                        'lookup_value_AUDIT' : ', '.join(set(filter(None, [z if pattern.search(z.replace('-', ' ').replace('.', ' ').strip()) is not None else "" for z in list_lookup_new]))),
                        'cell_ref_dcw' : str(e[0][1])
                        }
                    
                    output_list.append(mismatch_dict)
                #update counter
                x += 1

        writer_csv(output_list, j)
        
        
        #summary stats
        print '######## SUMMARY METRICS ########\n'
        print 'Refer to output file for full details: ' + "DCW_Compare_{dcw}.csv".format(dcw = j.rsplit('.',2)[0]) + '\n'
        for j in range(1, col_count + 1):
            counter = 0
            for i in output_list:
                if i['columns_compared'] == str(j):
                    counter += 1                    
            print 'Errors for ' + str(j) + ' columns compared: ' + str(counter) + '\n'
            
        
        
if __name__ == "__main__":
    sheet_checker(ready)
      
        
        


