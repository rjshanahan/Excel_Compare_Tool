#!/usr/bin/env python
# -*- coding: utf-8 -*-

#Richard Shanahan
#21 October 2015

import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl import load_workbook
import itertools
import hashlib
import csv


new = 'YOURNEWWORKBOOK.xlsx'
old = 'YOUROLDWORKBOOK.xlsx'


#function to generate MD5 checksum for file
def md5(fname):
    hash = hashlib.md5()
    with open(fname, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash.update(chunk)
    return hash.hexdigest()


    
#formatted message re: mismatched cells
def mismatch_printer(sheet, row_new_list, row_old_list):
    
    mismatches = [('OLD:', cell_new, 'NEW:',cell_old) for cell_new, cell_old in zip(row_new_list, row_old_list) if cell_new != cell_old]
    
    if len(mismatches) != 0:
        csv_writer(mismatches, sheet)
        #print 'SHEET: "' + sheet + '" has matching errors - refer to problem cell(s) below' + '\n\n'
        #print mismatches
        #rint '\n\n'
    else:
        pass
    
    
      
def csv_writer(mismatches, sheet):
          
    file_out = "COMPARE_{sheetname}_Errors.txt".format(sheetname = sheet)
    
    #with open(file_out, 'w') as csvfile:

    row1 = 'SHEET: "' + sheet + '" has matching errors - refer to problem cell(s) below' + '\n\n'
    row2 = mismatches
    row3 = '\n\n'
        
    text_file = open(file_out, "w")
    text_file.write(str(row1))
    text_file.write(str(row2))
    text_file.write(row3)
    text_file.close()
        
        
        
    
#iterate through sheets and identify cells that do not match 
def sheet_checker(new, old):    

    #load workbooks for DCW and Audit Report
    wb_new = openpyxl.load_workbook(new, use_iterators=True, data_only=True)
    wb_old = openpyxl.load_workbook(old, use_iterators=True, data_only=True)

    #extract sheet names
    sheet_list_new = wb_new.get_sheet_names()
    sheet_list_old = wb_old.get_sheet_names()
    

    for i, j in zip(sheet_list_new, sheet_list_old):

        ws_new = wb_new.get_sheet_by_name(i)
        ws_old = wb_old.get_sheet_by_name(j)

        row_new_list = []
        row_old_list = []
    
        for row_new, row_old in zip(ws_new.iter_rows(), ws_old.iter_rows()):
            [row_new_list.append([cell_new.coordinate, cell_new.internal_value]) for cell_new in row_new]
            [row_old_list.append([cell_old.coordinate, cell_old.internal_value]) for cell_old in row_old]
        
        if row_new_list != row_old_list:
            mismatch_printer(j, row_new_list, row_old_list)
        
        
                

    
#initiate file compares
if md5(new) != md5(old):
    print 'Oh dear... your files do not match'
    print 'new MD5 Checksum: ' + md5(new)
    print 'old MD5 Checksum: ' + md5(old) + '\n\n' + 'Check the Output Error files for details of the cells that don\'t match'
    
    if __name__ == "__main__":
        sheet_checker(new, old)
        
else:
    print 'Congratulations... your files match!'
    print 'new MD5 Checksum: ' + md5(new)
    print 'old MD5 Checksum: ' + md5(old)
