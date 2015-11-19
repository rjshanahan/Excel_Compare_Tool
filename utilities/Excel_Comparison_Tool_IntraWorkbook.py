#!/usr/bin/env python
# -*- coding: utf-8 -*-

#Richard Shanahan
#21 October 2015

## ADAPATED VERSION FOR CONSOLIDATED WORKBOOK ##

import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl import load_workbook
import itertools
import hashlib
import csv
import pprint as pp


ready = 'YOURWORKBOOK.xlsx'

sheet_list_old = ['list of "old" worksheet names to compare']
sheet_list_new = ['list of "new" worksheet names to compare']


#function to generate MD5 checksum for file
def md5(fname):
    hash = hashlib.md5()
    with open(fname, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash.update(chunk)
    return hash.hexdigest()


    
#formatted message re: mismatched cells
def mismatch_printer(sheet, row_new_list, row_old_list):
       
    mismatches = [('OLD:', cell_new, 'NEW:',cell_old) for cell_new, cell_old in zip(row_new_list, row_old_list) if cell_new != cell_old]
    
    if len(mismatches) != 0:
        txt_writer(mismatches, sheet)

    else:
        pass
    
    
#function to write mismatches to txt file
def txt_writer(mismatches, sheet):
    
    file_out = "COMPARE_{worksheet_compare}_Errors.txt".format(worksheet_compare = sheet)
        
    row1 = 'SHEET: "' + sheet + '" has matching errors - refer to problem cell(s) below' + '\n\n'
    row2 = pp.pformat(mismatches)
    row3 = '\n\n'
               
    text_file = open(file_out, "w")
    text_file.write(str(row1))
    text_file.write(str(row2))
    text_file.write(row3)
    text_file.close()
        
        
        
    
#iterate through sheets and identify cells that do not match 
#def sheet_checker(new, old):    
def sheet_checker(ready):    


    #load workbooks for DCW and Audit Report
    wb_all = openpyxl.load_workbook(ready, use_iterators=True, data_only=True)

 
    for i, j in zip(sheet_list_new, sheet_list_old):

        ws_new = wb_all.get_sheet_by_name(i)
        ws_old = wb_all.get_sheet_by_name(j)

        row_new_list = []
        row_old_list = []
    
        for row_new, row_old in zip(ws_new.iter_rows(), ws_old.iter_rows(row_offset=2)):
            [row_new_list.append([cell_new.coordinate, cell_new.internal_value]) for cell_new in row_new]
            [row_old_list.append([cell_old.coordinate, cell_old.internal_value]) for cell_old in row_old]
        
        if row_new_list != row_old_list:
            mismatch_printer(j, row_new_list, row_old_list)
            
    print 'mismatch results have been written to file - please review'

                
if __name__ == "__main__":
    sheet_checker(ready)
